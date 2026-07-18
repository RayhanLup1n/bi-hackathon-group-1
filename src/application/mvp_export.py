"""
Export service for MVP recommendations.

Generates CSV and Excel exports from recommendation dicts.
Uses stdlib csv for CSV and openpyxl (already in pyproject.toml) for Excel.
No new dependencies.

Functions:
  export_priorities_csv(priorities, reviews=None) → bytes
  export_priorities_xlsx(priorities, reviews=None) → bytes
  export_single_csv(detail, review=None) → bytes
  export_single_xlsx(detail, review=None) → bytes
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Shared constants ─────────────────────────────────────────────────────────

CSV_COLUMNS = [
    ("recommendation_id", "Recommendation ID"),
    ("rank", "Rank"),
    ("commodity", "Komoditas"),
    ("region", "Wilayah"),
    ("price_condition", "Kondisi Harga"),
    ("risk_level", "Tingkat Risiko"),
    ("display_priority_score", "Priority Score (Display)"),
    ("raw_priority_score", "Priority Score (Raw)"),
    ("confidence_factor", "Confidence Factor"),
    ("confidence_level", "Tingkat Keyakinan"),
    ("time_horizon_days", "Horizon (Hari)"),
    ("next_step", "Next Step"),
    ("knowledge_status", "Status Knowledge"),
]

# Neobrutalism-inspired header style
HEADER_FONT = Font(name="Inter", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2ECC88", end_color="2ECC88", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="medium", color="1A1A1A"),
    right=Side(style="medium", color="1A1A1A"),
    top=Side(style="medium", color="1A1A1A"),
    bottom=Side(style="medium", color="1A1A1A"),
)
DATA_FONT = Font(name="Inter", size=10)
DATA_ALIGNMENT = Alignment(vertical="center")

RISK_FILLS: dict[str, PatternFill] = {
    "rendah": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "sedang": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "tinggi": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "kritis": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
}


def _auto_width(worksheet) -> None:
    """Set column widths based on max content length in each column."""
    for col_cells in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                # Rough width: count chars, CJK chars count as ~2
                cell_str = str(cell.value)
                max_len = max(max_len, len(cell_str))
        worksheet.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _style_header(worksheet) -> None:
    """Apply neobrutalist header styling to first row."""
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _write_csv_rows(writer, priorities: list[dict[str, Any]], reviews: dict[str, Any] | None = None) -> None:
    """Write recommendation rows to a CSV writer."""
    reviews = reviews or {}
    for i, rec in enumerate(priorities, start=1):
        rid = rec.get("recommendation_id", "")
        review = reviews.get(rid, {})
        writer.writerow({
            "recommendation_id": rid,
            "rank": i,
            "commodity": rec.get("commodity", ""),
            "region": rec.get("region", ""),
            "price_condition": rec.get("price_condition", ""),
            "risk_level": rec.get("risk_level", ""),
            "display_priority_score": rec.get("display_priority_score", ""),
            "raw_priority_score": rec.get("raw_priority_score", ""),
            "confidence_factor": rec.get("confidence_factor", ""),
            "confidence_level": rec.get("confidence_level", ""),
            "time_horizon_days": rec.get("time_horizon_days", ""),
            "next_step": rec.get("next_step", ""),
            "knowledge_status": rec.get("knowledge_status", ""),
        })


# ── CSV exports ──────────────────────────────────────────────────────────────

def export_priorities_csv(
    priorities: list[dict[str, Any]],
    reviews: dict[str, Any] | None = None,
) -> bytes:
    """Export all priorities as CSV.

    Args:
        priorities: List of recommendation dicts from get_priorities().
        reviews: Optional dict mapping recommendation_id → review dict.

    Returns:
        UTF-8 encoded CSV bytes with BOM (for Excel compatibility).
    """
    output = io.StringIO(newline="")
    writer = csv.DictWriter(
        output,
        fieldnames=[col[0] for col in CSV_COLUMNS],
        extrasaction="ignore",
    )
    # Write header with display names
    writer.writerow(dict(CSV_COLUMNS))
    _write_csv_rows(writer, priorities, reviews)

    # BOM for Excel on Windows
    return ("﻿" + output.getvalue()).encode("utf-8")


def export_single_csv(detail: dict[str, Any], review: dict[str, Any] | None = None) -> bytes:
    """Export a single recommendation detail as CSV.

    Produces a two-section CSV:
      1. Summary block (key-value pairs)
      2. Evidence block (all evidence groups as rows)

    Args:
        detail: Full recommendation detail from get_priority_detail().
        review: Optional review dict.

    Returns:
        UTF-8 encoded CSV bytes with BOM.
    """
    output = io.StringIO(newline="")
    writer = csv.writer(output)

    # Section 1: Summary
    writer.writerow(["=== Ringkasan Rekomendasi ==="])
    summary_fields = [
        ("Komoditas", detail.get("commodity", "")),
        ("Wilayah", detail.get("region", "")),
        ("Kondisi Harga", detail.get("price_condition", "")),
        ("Tingkat Risiko", detail.get("risk_level", "")),
        ("Priority Score", detail.get("display_priority_score", "")),
        ("Confidence Level", detail.get("confidence_level", "")),
        ("Next Step", detail.get("next_step", "")),
        ("Horizon", f"{detail.get('time_horizon_days', '')} hari"),
    ]
    for key, val in summary_fields:
        writer.writerow([key, val])

    # Review status
    if review:
        writer.writerow([])
        writer.writerow(["Status Review", review.get("status", "")])
        writer.writerow(["Catatan Reviewer", review.get("note", "")])

    # Section 2: Evidence
    writer.writerow([])
    writer.writerow(["=== Evidence Groups ==="])
    writer.writerow(["Group", "Label", "Value"])

    for group_name, group_key in [
        ("Fakta Teramati", "observed_facts"),
        ("Output Model", "model_outputs"),
        ("Faktor Kemungkinan", "possible_factors"),
    ]:
        items = detail.get(group_key, []) or []
        for item in items:
            writer.writerow([
                group_name,
                item.get("label", ""),
                str(item.get("value", "")),
            ])

    # Missing information
    missing = detail.get("missing_information", []) or []
    for m in missing:
        writer.writerow(["Belum Tersedia", m, ""])

    # Sources
    writer.writerow([])
    writer.writerow(["=== Sumber Data ==="])
    writer.writerow(["Nama", "Cutoff"])
    for src in detail.get("sources", []) or []:
        writer.writerow([src.get("name", ""), src.get("cutoff", "")])

    return ("﻿" + output.getvalue()).encode("utf-8")


# ── Excel exports ────────────────────────────────────────────────────────────

def export_priorities_xlsx(
    priorities: list[dict[str, Any]],
    reviews: dict[str, Any] | None = None,
) -> bytes:
    """Export all priorities as styled Excel workbook.

    Creates 3 sheets:
      - "Prioritas": Flat table of all recommendations.
      - "Evidence": All evidence items across recommendations.
      - "Reviews": Human review data.

    Args:
        priorities: List of recommendation dicts.
        reviews: Optional dict mapping recommendation_id → review.

    Returns:
        XLSX file bytes.
    """
    reviews = reviews or {}
    wb = Workbook()

    # ── Sheet 1: Prioritas ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Prioritas"
    ws1.append([col[1] for col in CSV_COLUMNS])
    _style_header(ws1)

    for i, rec in enumerate(priorities, start=1):
        rid = rec.get("recommendation_id", "")
        row = [
            rid,
            i,
            rec.get("commodity", ""),
            rec.get("region", ""),
            rec.get("price_condition", ""),
            rec.get("risk_level", ""),
            rec.get("display_priority_score", ""),
            rec.get("raw_priority_score", ""),
            rec.get("confidence_factor", ""),
            rec.get("confidence_level", ""),
            rec.get("time_horizon_days", ""),
            rec.get("next_step", ""),
            rec.get("knowledge_status", ""),
        ]
        ws1.append(row)

        # Style data row
        risk = rec.get("risk_level", "")
        risk_fill = RISK_FILLS.get(risk)
        for col_idx in range(1, len(row) + 1):
            cell = ws1.cell(row=i + 1, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER
            # Color the risk level column
            if col_idx == 6 and risk_fill:
                cell.fill = risk_fill

    ws1.freeze_panes = "A2"
    _auto_width(ws1)

    # ── Sheet 2: Evidence ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Evidence")
    evidence_headers = ["Recommendation ID", "Komoditas", "Group", "Label", "Value"]
    ws2.append(evidence_headers)
    _style_header(ws2)

    for rec in priorities:
        rid = rec.get("recommendation_id", "")
        commodity = rec.get("commodity", "")
        for group_label, group_key in [
            ("Fakta Teramati", "observed_facts"),
            ("Output Model", "model_outputs"),
            ("Faktor Kemungkinan", "possible_factors"),
        ]:
            for item in rec.get(group_key, []) or []:
                ws2.append([
                    rid, commodity, group_label,
                    item.get("label", ""), str(item.get("value", "")),
                ])
        # Missing information
        for m in rec.get("missing_information", []) or []:
            ws2.append([rid, commodity, "Belum Tersedia", m, ""])

    ws2.freeze_panes = "A2"
    _auto_width(ws2)

    # ── Sheet 3: Reviews ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Reviews")
    review_headers = [
        "Recommendation ID", "Komoditas", "Status Review",
        "Reviewer ID", "Catatan", "Tanggal Review",
    ]
    ws3.append(review_headers)
    _style_header(ws3)

    for rec in priorities:
        rid = rec.get("recommendation_id", "")
        review = reviews.get(rid, rec.get("review"))
        if review:
            ws3.append([
                rid,
                rec.get("commodity", ""),
                review.get("status", ""),
                review.get("reviewer_user_id", ""),
                review.get("note", ""),
                review.get("updated_at", "") or review.get("created_at", ""),
            ])

    ws3.freeze_panes = "A2"
    _auto_width(ws3)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_single_xlsx(
    detail: dict[str, Any],
    review: dict[str, Any] | None = None,
) -> bytes:
    """Export a single recommendation as styled Excel workbook.

    Creates 4 sheets:
      - "Ringkasan": Key-value summary of the recommendation.
      - "Evidence": All evidence items.
      - "Response Options": Response options from rule engine.
      - "Sumber": Data sources.

    Args:
        detail: Full recommendation detail from get_priority_detail().
        review: Optional review dict.

    Returns:
        XLSX file bytes.
    """
    wb = Workbook()

    # ── Sheet 1: Ringkasan ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Ringkasan"
    ws1.append(["Field", "Value"])
    _style_header(ws1)

    summary_rows = [
        ("Komoditas", detail.get("commodity", "")),
        ("Wilayah", detail.get("region", "")),
        ("Kondisi Harga", detail.get("price_condition", "")),
        ("Tingkat Risiko", detail.get("risk_level", "")),
        ("Priority Score (Display)", detail.get("display_priority_score", "")),
        ("Priority Score (Raw)", detail.get("raw_priority_score", "")),
        ("Confidence Factor", detail.get("confidence_factor", "")),
        ("Confidence Level", detail.get("confidence_level", "")),
        ("Next Step", detail.get("next_step", "")),
        ("Time Horizon", f"{detail.get('time_horizon_days', '')} hari"),
        ("Knowledge Status", detail.get("knowledge_status", "")),
    ]

    for field, value in summary_rows:
        row_num = ws1.max_row + 1
        ws1.append([field, str(value)])
        for col in range(1, 3):
            cell = ws1.cell(row=row_num, column=col)
            cell.font = DATA_FONT

    if review:
        ws1.append([])
        ws1.append(["Status Review", review.get("status", "")])
        ws1.append(["Catatan Reviewer", review.get("note", "")])
        ws1.append(["Tanggal Review", review.get("updated_at", "") or review.get("created_at", "")])

    _auto_width(ws1)

    # ── Sheet 2: Evidence ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Evidence")
    ws2.append(["Group", "Label", "Value"])
    _style_header(ws2)

    for group_label, group_key in [
        ("Fakta Teramati", "observed_facts"),
        ("Output Model", "model_outputs"),
        ("Faktor Kemungkinan", "possible_factors"),
    ]:
        for item in detail.get(group_key, []) or []:
            ws2.append([group_label, item.get("label", ""), str(item.get("value", ""))])

    for m in detail.get("missing_information", []) or []:
        ws2.append(["Belum Tersedia", m, ""])

    _auto_width(ws2)

    # ── Sheet 3: Response Options ──────────────────────────────────────
    ws3 = wb.create_sheet("Response Options")
    ws3.append(["Tipe", "Label", "Deskripsi"])
    _style_header(ws3)
    for opt in detail.get("response_options", []) or []:
        ws3.append([opt.get("type", ""), opt.get("label", ""), opt.get("description", "")])
    _auto_width(ws3)

    # ── Sheet 4: Sumber ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Sumber")
    ws4.append(["Nama", "Cutoff", "Versi Model"])
    _style_header(ws4)
    for src in detail.get("sources", []) or []:
        ws4.append([
            src.get("name", ""),
            src.get("cutoff", ""),
            src.get("model_version", ""),
        ])
    _auto_width(ws4)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
