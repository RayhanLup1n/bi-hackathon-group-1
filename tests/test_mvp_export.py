"""
Unit tests for mvp_export.py — CSV and Excel export generation.

Tests verify that export functions produce valid output without errors.
No database, no HTTP — byte-level assertion tests.
"""
from __future__ import annotations

import csv
import io

import pytest
from openpyxl import load_workbook

from src.application.mvp_export import (
    export_priorities_csv,
    export_priorities_xlsx,
    export_single_csv,
    export_single_xlsx,
)


# ── Test fixtures ────────────────────────────────────────────────────────────

@pytest.fixture
def sample_priorities() -> list[dict]:
    """Minimal recommendation dicts matching orchestrator output."""
    return [
        {
            "recommendation_id": "rec_bawang_merah_banten",
            "commodity": "Bawang Merah",
            "region": "Nasional",
            "price_condition": "Di Bawah Ambang",
            "risk_level": "sedang",
            "display_priority_score": 45.0,
            "raw_priority_score": 60.0,
            "confidence_factor": 0.75,
            "confidence_level": "medium",
            "time_horizon_days": 7,
            "next_step": "Monitor harga bawang merah",
            "knowledge_status": "FACT",
            "observed_facts": [
                {"kind": "FACT", "label": "HET", "value": "75%"},
            ],
            "model_outputs": [],
            "possible_factors": [],
            "missing_information": ["Data stok tidak tersedia"],
            "response_options": [],
            "sources": [{"name": "PIHPS", "cutoff": "2026-07-13"}],
        },
        {
            "recommendation_id": "rec_cabai_rawit_jabar",
            "commodity": "Cabai Rawit Merah",
            "region": "Nasional",
            "price_condition": "Melampaui Ambang",
            "risk_level": "kritis",
            "display_priority_score": 92.5,
            "raw_priority_score": 98.0,
            "confidence_factor": 0.94,
            "confidence_level": "high",
            "time_horizon_days": 7,
            "next_step": "Verifikasi dan koordinasikan eskalsi",
            "knowledge_status": "MODEL_OUTPUT",
            "observed_facts": [
                {"kind": "FACT", "label": "HET", "value": "130%"},
            ],
            "model_outputs": [
                {"kind": "MODEL_OUTPUT", "label": "Forecast P90", "value": "Rp 95,000"},
            ],
            "possible_factors": [
                {"kind": "POSSIBLE_FACTOR", "label": "Cuaca", "value": "Hujan ekstrem"},
            ],
            "missing_information": ["Data stok cabai"],
            "response_options": [
                {"type": "VERIFIKASI", "label": "Verifikasi harga", "description": ""},
            ],
            "sources": [{"name": "PIHPS", "cutoff": "2026-07-13"}],
        },
    ]


@pytest.fixture
def single_detail() -> dict:
    """Minimal single recommendation detail."""
    return {
        "recommendation_id": "rec_cabai_rawit_jabar",
        "commodity": "Cabai Rawit Merah",
        "region": "Nasional",
        "price_condition": "Melampaui Ambang",
        "risk_level": "kritis",
        "display_priority_score": 92.5,
        "raw_priority_score": 98.0,
        "confidence_factor": 0.94,
        "confidence_level": "high",
        "time_horizon_days": 7,
        "next_step": "Verifikasi dan koordinasikan eskalsi",
        "knowledge_status": "MODEL_OUTPUT",
        "observed_facts": [
            {"kind": "FACT", "label": "HET", "value": "130%"},
            {"kind": "FACT", "label": "Harga Saat Ini", "value": "Rp 85,000/kg"},
        ],
        "model_outputs": [
            {"kind": "MODEL_OUTPUT", "label": "Forecast P90", "value": "Rp 95,000"},
        ],
        "possible_factors": [
            {"kind": "POSSIBLE_FACTOR", "label": "Cuaca", "value": "Hujan ekstrem"},
        ],
        "missing_information": ["Data stok cabai", "Volume distribusi"],
        "response_options": [
            {"type": "VERIFIKASI", "label": "Verifikasi harga", "description": "Cek sumber alternatif"},
            {"type": "KOORDINASIKAN", "label": "Koordinasi", "description": "Bahas dengan tim"},
        ],
        "sources": [
            {"name": "PIHPS", "cutoff": "2026-07-13", "model_version": ""},
        ],
    }


@pytest.fixture
def sample_reviews() -> dict:
    """Mock review dicts keyed by recommendation_id."""
    return {
        "rec_bawang_merah_banten": {
            "status": "Untuk Dibahas",
            "reviewer_user_id": 1,
            "note": "Perlu didiskusikan dalam rapat TPID",
            "created_at": "2026-07-13T10:00:00",
            "updated_at": "2026-07-13T10:30:00",
        },
    }


# ── CSV export tests ─────────────────────────────────────────────────────────

class TestCsvExport:
    """Verify CSV export output."""

    def test_priorities_csv_has_header_and_rows(self, sample_priorities):
        csv_bytes = export_priorities_csv(sample_priorities)
        content = csv_bytes.decode("utf-8")

        # BOM + header + 2 data rows = 4 lines minimum
        lines = content.strip().split("\n")
        assert len(lines) >= 3  # header + 2 rows

        # Verify header
        assert "Komoditas" in lines[0]
        assert "Tingkat Risiko" in lines[0]

    def test_priorities_csv_contains_commodity_names(self, sample_priorities):
        csv_bytes = export_priorities_csv(sample_priorities)
        content = csv_bytes.decode("utf-8")
        assert "Bawang Merah" in content
        assert "Cabai Rawit Merah" in content

    def test_priorities_csv_has_bom(self, sample_priorities):
        csv_bytes = export_priorities_csv(sample_priorities)
        assert csv_bytes[:3] == b"\xef\xbb\xbf"

    def test_empty_priorities_returns_header_only(self):
        csv_bytes = export_priorities_csv([])
        content = csv_bytes.decode("utf-8")
        lines = content.strip().split("\n")
        # Header only, no data rows
        assert len(lines) == 1
        assert "Komoditas" in lines[0]

    def test_priorities_csv_valid_utf8_roundtrip(self, sample_priorities):
        csv_bytes = export_priorities_csv(sample_priorities)
        # Decode then re-parse with csv reader — header uses display names
        reader = csv.DictReader(io.StringIO(csv_bytes.decode("utf-8")))
        rows = list(reader)
        assert len(rows) == 2
        assert rows[0]["Komoditas"] == "Bawang Merah"
        assert rows[1]["Komoditas"] == "Cabai Rawit Merah"

    def test_single_csv_has_sections(self, single_detail):
        csv_bytes = export_single_csv(single_detail)
        content = csv_bytes.decode("utf-8")
        assert "Ringkasan Rekomendasi" in content
        assert "Evidence Groups" in content
        assert "Sumber Data" in content
        assert "Cabai Rawit Merah" in content

    def test_single_csv_includes_evidence_items(self, single_detail):
        csv_bytes = export_single_csv(single_detail)
        content = csv_bytes.decode("utf-8")
        assert "Fakta Teramati" in content
        assert "Output Model" in content
        assert "Faktor Kemungkinan" in content
        assert "Belum Tersedia" in content

    def test_single_csv_with_review(self, single_detail):
        review = {"status": "Untuk Dibahas", "note": "Segera bahas"}
        csv_bytes = export_single_csv(single_detail, review=review)
        content = csv_bytes.decode("utf-8")
        assert "Untuk Dibahas" in content
        assert "Segera bahas" in content


# ── Excel export tests ───────────────────────────────────────────────────────

class TestXlsxExport:
    """Verify Excel export output."""

    def test_priorities_xlsx_has_three_sheets(self, sample_priorities, sample_reviews):
        xlsx_bytes = export_priorities_xlsx(sample_priorities, sample_reviews)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_names = wb.sheetnames
        assert "Prioritas" in sheet_names
        assert "Evidence" in sheet_names
        assert "Reviews" in sheet_names

    def test_priorities_xlsx_first_sheet_has_data(self, sample_priorities):
        xlsx_bytes = export_priorities_xlsx(sample_priorities)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Prioritas"]
        # Header + 2 data rows
        assert ws.max_row == 3
        assert ws.cell(1, 3).value == "Komoditas"  # header column 3

    def test_priorities_xlsx_evidence_sheet_has_rows(self, sample_priorities):
        xlsx_bytes = export_priorities_xlsx(sample_priorities)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Evidence"]
        assert ws.max_row >= 2  # header + at least 1 evidence row

    def test_priorities_xlsx_header_is_styled(self, sample_priorities):
        xlsx_bytes = export_priorities_xlsx(sample_priorities)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Prioritas"]
        header_cell = ws.cell(1, 1)
        assert header_cell.font.bold is True

    def test_empty_priorities_xlsx_still_creates_sheets(self):
        xlsx_bytes = export_priorities_xlsx([])
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "Prioritas" in wb.sheetnames
        assert "Evidence" in wb.sheetnames
        assert "Reviews" in wb.sheetnames

    def test_single_xlsx_has_four_sheets(self, single_detail):
        xlsx_bytes = export_single_xlsx(single_detail)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_names = wb.sheetnames
        assert "Ringkasan" in sheet_names
        assert "Evidence" in sheet_names
        assert "Response Options" in sheet_names
        assert "Sumber" in sheet_names

    def test_single_xlsx_ringkasan_has_summary_data(self, single_detail):
        xlsx_bytes = export_single_xlsx(single_detail)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Ringkasan"]
        assert ws.cell(2, 1).value == "Komoditas"
        assert ws.cell(2, 2).value == "Cabai Rawit Merah"

    def test_single_xlsx_opens_without_error(self, single_detail):
        xlsx_bytes = export_single_xlsx(single_detail)
        # load_workbook succeeds = file is valid
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert wb is not None
